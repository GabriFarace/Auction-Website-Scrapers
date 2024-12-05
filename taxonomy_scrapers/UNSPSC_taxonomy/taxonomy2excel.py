import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def get_table(taxonomy):

    # Flatten the taxonomy
    flattened_data = []

    def traverse(node, path):
        category_name = node.get('title', '')
        category_code = node.get('code', '')
        flattened_data.append([category_code] + path + [category_name])
        for subcategory in node.get('subcategories', []):
            traverse(subcategory, path + [category_name])

    # Start traversal
    for category in taxonomy:
        traverse(category, [])

    # Determine the maximum depth of the hierarchy
    max_depth = 5  # Maximum depth set to 5
    # cat, l1, l2, l3, l4, l5, l6
    #   0,  1,  2,  3,  4,  5,  6
    # Pad rows to ensure they match the maximum depth
    for row in flattened_data:
        if len(row) < max_depth + 1:
            while len(row) < max_depth + 1:
                row.insert(len(row), '')
        '''elif len(row) > max_depth + 1:
            sub_list = row[max_depth:]
            new_cat_name = "/".join(sub_list)
            while len(row) > max_depth:
                row.pop(max_depth)
            row.append(new_cat_name)'''
    flattened_data_max = [row for row in flattened_data if len(row) == max_depth + 1]

    # Define column names for levels and attributes
    level_columns = [f'Level {i} Name' for i in range(1, max_depth + 1)]
    attribute_columns = [
        'useful_life_years', 'useful_life_kwh', 'useful_life_hours',
        'ratio_energy_hours', 'useful_life_unit_produced', 'ratio_energy_unit',
        'obsolescence_coefficient', 'residual_value'
    ]
    # Create DataFrame and add empty attribute columns
    df = pd.DataFrame(flattened_data_max, columns=["Category code"] + level_columns)
    for attr in attribute_columns:
        df[attr] = ''

    return df


def group_and_color(df, excel_name_input, excel_name_output):

    # Load the workbook for grouping and formatting
    workbook = load_workbook(excel_name_input)
    sheet = workbook.active
    sheet.title = 'Table'

    # Define colors for each level
    colors = [
        "FFDAB9",  # Level 2: Peach
        "FFFACD",  # Level 3: Lemon Chiffon
        "E6E6FA",  # Level 4: Lavender
        "D3FFCE"  # Level 5: Light Green
    ]


    # Programmatically group rows based on hierarchy depth
    def build_group(rows_list, base, depth):
        if depth == 5:
            return
        check = False
        current_ind = 0
        row_ind = 1
        for i, row_element in enumerate(rows_list):
            row_ind = i

            if row_element[depth + 1] == '':# Check for the first non-empty level in the hierarchy
                fill_color = colors[depth - 1]
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                for cell in sheet[base + row_ind + 2]:  # Apply fill to all cells in the row
                    cell.fill = fill
                if check:
                    if not row_ind == current_ind + 1:
                        sheet.row_dimensions.group(start=base + current_ind + 1 + 2, end= base + row_ind - 1 + 2, outline_level=depth)
                        print(f"LEVEL {depth}:Grouping from index{base + current_ind + 1 + 2} to index{base + row_ind - 1 + 2}.")
                        build_group(rows_list[current_ind + 1 : row_ind], base + current_ind + 1, depth + 1)
                    current_ind = row_ind
                else:
                    check = True
                    current_ind = row_ind

        if not row_ind == current_ind:
            print(f"OUTER LEVEL {depth}:Grouping from index{base + current_ind + 1 + 2} to index{base + row_ind + 2}.")
            sheet.row_dimensions.group(start= base + current_ind  + 1 + 2, end=base + row_ind + 2, outline_level=depth)
            build_group(rows_list[current_ind + 1: row_ind + 1], base + current_ind + 1, depth + 1)

    rows_as_list = df.values.tolist()
    build_group(rows_as_list, 0, 1)
    # Auto-resize columns to fit content
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Save the final grouped and formatted Excel file
    final_excel_filename = excel_name_output
    workbook.save(final_excel_filename)

    print(f"Excel file created and saved as '{final_excel_filename}'.")

def create_legend(excel_name_input, excel_name_output):

    # Load the existing workbook
    workbook = load_workbook(excel_name_input)

    # Create a new sheet for the Legend
    legend_sheet = workbook.create_sheet(title="Legend")

    # Define attribute explanations
    attributes = [
        ("Category code", "Code of the category"),
        ("Level 1 Name", "Top-level categories"),
        ("Level 2 Name", "Subcategories under Level 1"),
        ("Level 3 Name", "Subcategories under Level 2"),
        ("Level 4 Name", "Subcategories under Level 3"),
        ("Level 5 Name", "Leaves level or specific items"),
        ("useful_life_years", "Expected useful life of the asset in years."),
        ("useful_life_kwh", "Expected useful life in terms of energy consumption (kWh)."),
        ("useful_life_hours", "Expected useful life in total operating hours."),
        ("ratio_energy_hours", "Ratio of energy consumed per hour of operation."),
        ("useful_life_unit_produced", "Total units produced over the asset's lifetime."),
        ("ratio_energy_unit", "Ratio of energy consumed per unit produced."),
        ("obsolescence_coefficient", "Coefficient representing percentage in annual loss of the asset value due to obsolescence"),
        ("residual_value", "Estimated value at the end of the asset's useful life.")
    ]




    # Add a title to the legend
    legend_sheet["A1"] = "Attribute Descriptions"
    legend_sheet["A1"].font = Font(bold=True, size=14)
    legend_sheet["A1"].alignment = Alignment(horizontal="center")

    # Write attribute explanations to the Legend sheet
    legend_sheet.append(["Attribute", "Description"])
    for attr, desc in attributes:
        legend_sheet.append([attr, desc])

    # Format the attribute headers
    for cell in legend_sheet["A2:B2"]:
        for c in cell:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

    #     # Apply colors to the "Color" column cells in the legend

    # Define colors for each level
    color_fills = [
        "FFDAB9",  # Level 1: Peach
        "FFFACD",  # Level 2: Lemon Chiffon
        "E6E6FA",  # Level 3: Lavender
        "D3FFCE",  # Level 4: Light Green
        "FFFFFF",  # Level 5: White
    ]

    for i, hex_color in enumerate(color_fills, start=4):  # Start at row 7 (after header)
        cell = legend_sheet[f"B{i}"]
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
        cell.fill = fill

    # Auto-adjust column widths for better readability
    for col in range(1, 3):  # Columns A to C
        max_length = 0
        col_letter = get_column_letter(col)
        for cell in legend_sheet[col_letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        legend_sheet.column_dimensions[col_letter].width = max_length + 5

    # Save the updated workbook
    workbook.save(excel_name_output)

    print(f"Excel file with legend created and saved as '{excel_name_output}'.")


def build_excel(taxonomy_path, output_file_name):
    with open(taxonomy_path, 'r') as f:
        taxonomy_list = json.load(f)
        #taxonomy_list = taxonomy_list[1:-1]
    df = get_table(taxonomy_list)

    # Save to Excel
    excel_intermediate1_filename = 'intermediate_1.xlsx'
    df.to_excel(excel_intermediate1_filename, index=False)

    excel_intermediate2_filename = 'intermediate_2.xlsx'
    group_and_color(df, excel_intermediate1_filename, excel_intermediate2_filename)
    create_legend(excel_intermediate2_filename, output_file_name)



if __name__ == '__main__':
    build_excel("taxonomy.json", "UNSPSC_taxonomy.xlsx")